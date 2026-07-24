#!/usr/bin/env python3
"""
pdb_csv_to_motif_and_enzyme.py

One pipeline, two analyses. From a CSV of PDB IDs, fetch each sequence ONCE
from RCSB, then run BOTH:

  (A) MOTIF scan  -- every sequence vs the full PROSITE database via
      Bio.ExPASy.ScanProsite  (finds PHD, MYND, SET, RING, ... signatures)
  (B) ENZYME ID   -- every sequence vs RCSB's MMseqs2 sequence-similarity
      service ("BLAST-like"), with top hits enriched to molecule name + EC
      number via the RCSB Data API.

Combines pdb_csv_to_motif_summary.py and pdb_csv_to_enzyme_id.py so the RCSB
FASTA fetch, the CSV parsing/grouping, and the intermediate FASTA files are
shared instead of done twice. Output is a SINGLE workbook with four sheets:

    motif_summary   -- signatures found per group, seq counts / %
    motif_hits      -- one row per PROSITE hit
    enzyme_summary  -- one row per query sequence: best hit + consensus name/EC
    enzyme_hits     -- one row per (query, sequence-search hit)

Either analysis can be turned off:
    --skip-motif    run only the enzyme-ID search
    --skip-enzyme   run only the PROSITE motif scan

Requires:  pip install biopython requests pandas openpyxl
Needs network access to www.rcsb.org, prosite.expasy.org, search.rcsb.org,
data.rcsb.org.

Usage:
    python pdb_csv_to_motif_and_enzyme.py INPUT.csv -o results.xlsx
    python pdb_csv_to_motif_and_enzyme.py INPUT.csv --id-col pdb_id \
        --group-col cluster --keep-fasta ./fastas
    python pdb_csv_to_motif_and_enzyme.py INPUT.csv --skip-enzyme    # motif only
    python pdb_csv_to_motif_and_enzyme.py INPUT.csv --skip-motif \
        --identity 0.3 --evalue 1 --top 10                           # enzyme only

Notes:
    - ScanProsite: keep --chunk small (<=10). --lowscore adds weak profile hits.
    - Enzyme search: --identity is the MMseqs2 cutoff (0-1); lower finds remote
      homologs. Since queries ARE PDB sequences they self-hit at 100%; the
      best NON-self hit and a consensus across non-self hits are reported.
    - Sequences are deduped and cached, so identical chains across groups are
      fetched, scanned, and searched only once.
"""

import argparse
import json
import os
import sys
import time
from collections import defaultdict, OrderedDict, Counter

import requests

# --------------------------------------------------------------------------- #
# Endpoints / tunables
# --------------------------------------------------------------------------- #
FASTA_URL = "https://www.rcsb.org/fasta/entry/{ids}"
SEARCH_URL = "https://search.rcsb.org/rcsbsearch/v2/query"
DATA_GRAPHQL_URL = "https://data.rcsb.org/graphql"

BATCH_SIZE = 50          # IDs / entity-ids per HTTP request
PAUSE = 0.3              # polite delay between RCSB fetch requests
MAX_RETRIES = 4
TIMEOUT = 60
MIN_SEQ_LEN = 25         # RCSB sequence search requires >=25 residues


# ========================================================================== #
# SHARED: RCSB FASTA fetch + CSV grouping
# ========================================================================== #
def fetch_batch(batch):
    """Return raw FASTA text for a comma-separated batch, or None on failure."""
    url = FASTA_URL.format(ids=",".join(batch))
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.get(url, timeout=TIMEOUT)
            if r.status_code == 200 and r.text.startswith(">"):
                return r.text
            if r.status_code == 404:
                return None
        except requests.RequestException:
            pass
        time.sleep(PAUSE * attempt * 2)
    return None


def parse_fasta_text(text):
    header, chunks = None, []
    for line in text.splitlines():
        if line.startswith(">"):
            if header is not None:
                yield header, "".join(chunks)
            header, chunks = line[1:], []
        else:
            chunks.append(line.strip())
    if header is not None:
        yield header, "".join(chunks)


def pdbid_from_header(header):
    return header.split("_", 1)[0].split("|", 1)[0].strip().upper()


def fetch_ids(ids, cache):
    """Fetch FASTA for `ids`, using/filling `cache` (pid -> [(hdr, seq), ...]).

    Returns (raw_records, got_ids_set). raw_records preserve input order.
    """
    need = [pid for pid in ids if pid not in cache]
    for i in range(0, len(need), BATCH_SIZE):
        batch = need[i:i + BATCH_SIZE]
        text = fetch_batch(batch)
        chunk_texts = []
        if text is None:
            for pid in batch:
                t = fetch_batch([pid])
                if t:
                    chunk_texts.append(t)
        else:
            chunk_texts.append(text)
        for t in chunk_texts:
            for header, seq in parse_fasta_text(t):
                if not seq:
                    continue
                pid = pdbid_from_header(header)
                cache.setdefault(pid, []).append((header, seq))
        for pid in batch:
            cache.setdefault(pid, [])              # mark tried-but-empty
        if need:
            done = min(i + BATCH_SIZE, len(need))
            print(f"    fetched {done}/{len(need)} new IDs ...", flush=True)
            time.sleep(PAUSE)
    got, records = set(), []
    for pid in ids:
        for header, seq in cache.get(pid, []):
            got.add(pid)
            records.append((pid, header, seq))
    return records, got


def build_records(raw_records, all_chains):
    """Dedup per group and format headers. raw_records: (pid, header, seq).

    Returns OrderedDict: key -> (entity_id, seq).
    """
    out = OrderedDict()
    for pid, header, seq in raw_records:
        key = header if all_chains else (pid, seq)
        if key not in out:
            short = header.split("|", 1)[0]        # e.g. "3M5L_1"
            out[key] = (short, seq)
    return out


def read_groups(path, id_col, group_col, no_group):
    """Return OrderedDict: group_name -> ordered list of unique PDB IDs."""
    import pandas as pd
    df = pd.read_csv(path, dtype=str)
    df.columns = [c.strip() for c in df.columns]

    id_col_actual = id_col if id_col in df.columns else df.columns[0]
    if id_col not in df.columns:
        print(f"  [warn] id column '{id_col}' not found; using "
              f"'{id_col_actual}'")

    if no_group or group_col not in df.columns:
        if not no_group and group_col not in df.columns:
            print(f"  [warn] group column '{group_col}' not found; "
                  f"scanning all IDs as one group")
        df = df.assign(_grp="all")
        group_col_actual = "_grp"
    else:
        group_col_actual = group_col

    groups = OrderedDict()
    for _, row in df.iterrows():
        raw = row.get(id_col_actual)
        if raw is None or str(raw).strip() == "" or str(raw).lower() == "nan":
            continue
        pid = str(raw).strip().upper()
        grp = str(row.get(group_col_actual, "all")).strip()
        groups.setdefault(grp, [])
        if pid not in groups[grp]:
            groups[grp].append(pid)
    return groups


def write_fasta_file(records, out_path):
    with open(out_path, "w") as fh:
        for sid, seq in records.values():
            fh.write(">" + sid + "\n")
            for j in range(0, len(seq), 60):
                fh.write(seq[j:j + 60] + "\n")


# ========================================================================== #
# (A) MOTIF: ScanProsite
# ========================================================================== #
NAME_MAP = {
    "PS50016": "ZF_PHD_2", "PS01359": "ZF_PHD_1", "PS50280": "SET",
    "PS50867": "PRESET", "PS50868": "POSTSET", "PS01360": "ZF_MYND_1",
    "PS50865": "ZF_MYND_2", "PS00028": "ZINC_FINGER_C2H2_1",
    "PS50157": "ZINC_FINGER_C2H2_2", "PS00518": "ZF_RING_1",
    "PS50089": "ZF_RING_2", "PS00190": "CYTOCHROME_C",
    "PS00141": "ZINC_PROTEASE",
}


def chunks(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def load_prosite_names(path):
    """Parse ID/AC lines from prosite.dat -> {accession: name}. Optional."""
    names, cur_id = {}, None
    with open(path, errors="ignore") as fh:
        for line in fh:
            if line.startswith("ID "):
                cur_id = line[5:].split(";")[0].strip()
            elif line.startswith("AC "):
                ac = line[5:].strip().rstrip(";")
                if cur_id:
                    names[ac] = cur_id
    return names


def name_for(ac, names):
    return names.get(ac) or NAME_MAP.get(ac, ac)


def substring(seq, start, stop):
    if not seq or start is None or stop is None:
        return None
    s = max(1, start) - 1
    e = min(len(seq), stop)
    return seq[s:e] if s < e else None


def scan_block(block, lowscore, max_retries, sleep):
    """block = list of (id, seq). Returns list of Biopython match dicts."""
    from Bio.ExPASy import ScanProsite
    fasta = "\n".join(f">{sid}\n{seq}" for sid, seq in block)
    kw = {"lowscore": 1} if lowscore else {}
    for attempt in range(1, max_retries + 1):
        try:
            handle = ScanProsite.scan(seq=fasta, **kw)
            return list(ScanProsite.read(handle))
        except Exception as e:                                    # noqa: BLE001
            if attempt == max_retries:
                sys.stderr.write(f"  block failed after {attempt} tries: {e}\n")
                return []
            wait = sleep * attempt
            sys.stderr.write(f"  retry {attempt} in {wait:.0f}s ({e})\n")
            time.sleep(wait)
    return []


def motif_scan_group(records, lowscore, chunk, max_retries, sleep):
    """Run ScanProsite over one group's records. Returns raw match list."""
    nblocks = (len(records) + chunk - 1) // chunk
    matches = []
    for bi, block in enumerate(chunks(records, chunk), 1):
        print(f"  [motif] scanprosite block {bi}/{nblocks} ...", flush=True)
        matches.extend(scan_block(block, lowscore, max_retries, sleep))
        time.sleep(sleep)
    return matches


def motif_summarize(group_name, records, matches, names):
    """matches -> (summary_rows, detail_rows) for one group."""
    n = len(records)
    seqmap = dict(records)
    seqs_with = defaultdict(set)
    detail = []
    for m in matches:
        ac = m.get("signature_ac")
        if not ac:
            continue
        sid = m.get("sequence_id") or m.get("sequence_ac")
        nm = name_for(ac, names)
        start, stop = m.get("start"), m.get("stop")
        sub = substring(seqmap.get(sid, ""), start, stop)
        seqs_with[(ac, nm)].add(sid)
        detail.append([group_name, sid, ac, nm, start, stop, sub])

    summary = []
    if not seqs_with:
        summary.append([group_name, n, "(none)", "(none)", 0, "0.0%"])
    for (ac, nm), sset in sorted(seqs_with.items(), key=lambda kv: -len(kv[1])):
        c = len(sset)
        pct = 100.0 * c / n if n else 0
        summary.append([group_name, n, ac, nm, c, f"{pct:.1f}%"])
    return summary, detail


# ========================================================================== #
# (B) ENZYME: RCSB sequence search + Data-API enrichment
# ========================================================================== #
def sequence_search(seq, identity_cutoff, evalue_cutoff, rows, max_retries,
                    sleep):
    """One MMseqs2 sequence search. Returns list of hit dicts:
        {entity_id, score, identity(0-1), evalue, bitscore}.
    """
    payload = {
        "query": {
            "type": "terminal",
            "service": "sequence",
            "parameters": {
                "evalue_cutoff": evalue_cutoff,
                "identity_cutoff": identity_cutoff,
                "sequence_type": "protein",
                "value": seq,
            },
        },
        "return_type": "polymer_entity",
        "request_options": {
            "scoring_strategy": "sequence",
            "paginate": {"start": 0, "rows": rows},
            "results_verbosity": "verbose",
        },
    }
    for attempt in range(1, max_retries + 1):
        try:
            r = requests.post(SEARCH_URL, json=payload, timeout=TIMEOUT)
            if r.status_code == 204:
                return []
            if r.status_code == 200:
                return _parse_search_hits(r.json())
            if 400 <= r.status_code < 500:
                sys.stderr.write(
                    f"  search HTTP {r.status_code}: {r.text[:200]}\n")
                return []
        except requests.RequestException as e:
            if attempt == max_retries:
                sys.stderr.write(f"  search failed: {e}\n")
                return []
        except json.JSONDecodeError:
            return []
        time.sleep(sleep * attempt)
    return []


def _parse_search_hits(data):
    hits = []
    for item in data.get("result_set", []):
        entity_id = item.get("identifier")
        ident = evalue = bitscore = None
        best_score = item.get("score")
        for svc in item.get("services", []):
            for node in svc.get("nodes", []):
                for mc in node.get("match_context", []):
                    ident = mc.get("sequence_identity", ident)
                    evalue = mc.get("evalue", evalue)
                    bitscore = mc.get("bitscore", bitscore)
        hits.append({
            "entity_id": entity_id,
            "score": best_score,
            "identity": ident,
            "evalue": evalue,
            "bitscore": bitscore,
        })
    return hits


ENRICH_QUERY = """
query($ids: [String!]!) {
  polymer_entities(entity_ids: $ids) {
    rcsb_id
    rcsb_polymer_entity {
      pdbx_description
      rcsb_ec_lineage { id }
    }
    rcsb_polymer_entity_container_identifiers {
      entry_id
      reference_sequence_identifiers {
        database_name
        database_accession
      }
    }
  }
}
"""


def enrich_entities(entity_ids, cache, sleep=0.2):
    """Fill cache[entity_id] = {name, ec, entry}. Batched GraphQL."""
    need = [e for e in entity_ids if e not in cache]
    for i in range(0, len(need), BATCH_SIZE):
        batch = need[i:i + BATCH_SIZE]
        body = {"query": ENRICH_QUERY, "variables": {"ids": batch}}
        result = None
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                r = requests.post(DATA_GRAPHQL_URL, json=body, timeout=TIMEOUT)
                if r.status_code == 200:
                    result = r.json()
                    break
            except requests.RequestException:
                pass
            time.sleep(sleep * attempt * 2)
        if result:
            for ent in (result.get("data", {})
                        .get("polymer_entities", []) or []):
                if not ent:
                    continue
                eid = ent.get("rcsb_id")
                pe = ent.get("rcsb_polymer_entity") or {}
                name = pe.get("pdbx_description")
                ec_list = pe.get("rcsb_ec_lineage") or []
                ecs = [x.get("id") for x in ec_list if x.get("id")]
                full_ec = max(ecs, key=lambda s: s.count("."), default=None) \
                    if ecs else None
                ident = (ent.get("rcsb_polymer_entity_container_identifiers")
                         or {})
                refs = ident.get("reference_sequence_identifiers") or []
                uniprots = []
                for ref in refs:
                    if not ref:
                        continue
                    dbn = (ref.get("database_name") or "").upper()
                    acc = ref.get("database_accession")
                    if acc and dbn in ("UNIPROT", "UNIPROTKB"):
                        if acc not in uniprots:
                            uniprots.append(acc)
                cache[eid] = {"name": name, "ec": full_ec,
                              "entry": ident.get("entry_id"),
                              "uniprot": ";".join(uniprots) or None}
        for eid in batch:
            cache.setdefault(eid, {"name": None, "ec": None,
                                   "entry": None, "uniprot": None})
        time.sleep(sleep)
    return cache


def _majority(counter):
    if not counter:
        return None, 0.0
    total = sum(counter.values())
    val, cnt = counter.most_common(1)[0]
    return val, cnt / total


def _fmt_pct(x):
    if x is None:
        return None
    try:
        return f"{float(x) * 100:.1f}%" if float(x) <= 1 else f"{float(x):.1f}%"
    except (TypeError, ValueError):
        return None


def _fmt_eval(x):
    if x is None:
        return None
    try:
        return f"{float(x):.2e}"
    except (TypeError, ValueError):
        return None


def _fmt_num(x):
    if x is None:
        return None
    try:
        return round(float(x), 1)
    except (TypeError, ValueError):
        return None


def identify_sequence(group_name, query_id, seq, hits, enrich_cache, top):
    """hits for one query -> (summary_row, detail_rows), both group-prefixed."""
    query_entry = query_id.split("_", 1)[0].upper()
    top_hits = hits[:top]

    detail = []
    ec_votes, name_votes, uniprot_votes = Counter(), Counter(), Counter()
    best_other = None

    for h in top_hits:
        eid = h["entity_id"]
        meta = enrich_cache.get(eid, {})
        name, ec = meta.get("name"), meta.get("ec")
        uniprot = meta.get("uniprot")
        is_self = (eid.split("_", 1)[0].upper() == query_entry)
        detail.append([
            group_name, query_id, eid, "self" if is_self else "hit",
            name, ec, uniprot,
            _fmt_pct(h.get("identity")),
            _fmt_eval(h.get("evalue")),
            _fmt_num(h.get("bitscore")),
        ])
        if not is_self:
            if ec:
                ec_votes[ec] += 1
            if name:
                name_votes[name.strip()] += 1
            if uniprot:
                uniprot_votes[uniprot] += 1
            if best_other is None:
                best_other = (eid, name, ec, uniprot, h)

    cons_ec, ec_frac = _majority(ec_votes)
    cons_name, _ = _majority(name_votes)
    cons_uniprot, _ = _majority(uniprot_votes)

    if best_other is not None:
        b_eid, b_name, b_ec, b_uniprot, b_h = best_other
        best_id, best_name, best_ec = b_eid, b_name, b_ec
        best_uniprot = b_uniprot
        best_ident = _fmt_pct(b_h.get("identity"))
        best_eval = _fmt_eval(b_h.get("evalue"))
        best_bits = _fmt_num(b_h.get("bitscore"))
    else:
        best_id = best_name = best_ec = best_uniprot = None
        best_ident = best_eval = best_bits = None

    summary = [
        group_name, query_id, len(seq),
        best_id, best_name, best_ec, best_uniprot,
        best_ident, best_eval, best_bits,
        cons_name, cons_ec, cons_uniprot,
        f"{ec_frac:.0%}" if cons_ec else "0%",
        len(top_hits),
    ]
    return summary, detail


def enzyme_id_group(group_name, records, identity, evalue, rows, top,
                    max_retries, sleep, search_cache, enrich_cache):
    """Run enzyme-ID search over one group's records.

    Returns (summary_rows, detail_rows).
    """
    per_query_hits = {}
    for qi, (qid, seq) in enumerate(records, 1):
        if len(seq) < MIN_SEQ_LEN:
            print(f"  [enzyme] skip {qid}: sequence < {MIN_SEQ_LEN} aa")
            per_query_hits[qid] = (seq, [])
            continue
        if seq in search_cache:
            hits = search_cache[seq]
        else:
            print(f"  [enzyme] search {qi}/{len(records)}  {qid} ...",
                  flush=True)
            hits = sequence_search(seq, identity, evalue, rows,
                                   max_retries, sleep)
            search_cache[seq] = hits
            time.sleep(sleep)
        per_query_hits[qid] = (seq, hits)

    wanted = set()
    for qid, (seq, hits) in per_query_hits.items():
        for h in hits[:top]:
            if h.get("entity_id"):
                wanted.add(h["entity_id"])
    if wanted:
        print(f"  [enzyme] enriching {len(wanted)} hit entities ...",
              flush=True)
        enrich_entities(sorted(wanted), enrich_cache)

    summary_rows, detail_rows = [], []
    for qid, (seq, hits) in per_query_hits.items():
        s, d = identify_sequence(group_name, qid, seq, hits,
                                 enrich_cache, top)
        summary_rows.append(s)
        detail_rows.extend(d)
        if s[4]:
            print(f"    {qid:8s} -> {s[4]}"
                  + (f"  [EC {s[5]}]" if s[5] else ""))
    return summary_rows, detail_rows


# ========================================================================== #
# Excel output (single workbook, up to four sheets)
# ========================================================================== #
def build_overview(motif_detail, enzyme_summary, do_motif, do_enzyme,
                   n_motif_cols=3):
    """Join motif + enzyme results into one row per PDB ID.

    Columns: group, pdb_id, best_hit_name, consensus_name, motif1..motifN.
    - PDB ID is the 4-char entry (entities like 3M5L_1 collapse to 3M5L).
    - Enzyme names: first non-empty best/consensus name seen across the
      entry's entities.
    - Motifs: the entry's distinct PROSITE signature names, ordered by
      overall prevalence (most common across ALL entries first), spread
      across motif1..motifN. Extra motifs beyond N are joined into the
      last column so nothing is silently dropped.
    """
    def entry_of(entity):
        return str(entity).split("_", 1)[0].upper()

    # --- motif signatures per PDB entry, plus global prevalence ---
    motifs_by_pdb = defaultdict(set)         # pdb -> {signature, ...}
    group_by_pdb = {}                        # pdb -> group (first seen)
    global_motif_counts = Counter()          # signature -> n entries
    if do_motif:
        # count each signature once per entry for prevalence ordering
        seen_pair = set()
        for row in motif_detail:
            grp, sid, ac, sig = row[0], row[1], row[2], row[3]
            if sig in ("(none)", None):
                continue
            pdb = entry_of(sid)
            group_by_pdb.setdefault(pdb, grp)
            motifs_by_pdb[pdb].add(sig)
            if (pdb, sig) not in seen_pair:
                seen_pair.add((pdb, sig))
                global_motif_counts[sig] += 1

    # --- enzyme names per PDB entry ---
    best_by_pdb, cons_by_pdb = {}, {}
    if do_enzyme:
        for row in enzyme_summary:
            grp, qentity = row[0], row[1]
            best_name, cons_name = row[4], row[10]
            pdb = entry_of(qentity)
            group_by_pdb.setdefault(pdb, grp)
            if best_name and pdb not in best_by_pdb:
                best_by_pdb[pdb] = best_name
            if cons_name and pdb not in cons_by_pdb:
                cons_by_pdb[pdb] = cons_name

    # global ordering of signatures: most prevalent first, then name
    motif_order = [sig for sig, _ in sorted(
        global_motif_counts.items(), key=lambda kv: (-kv[1], kv[0]))]
    rank = {sig: i for i, sig in enumerate(motif_order)}

    all_pdbs = sorted(set(group_by_pdb))
    rows = []
    for pdb in all_pdbs:
        grp = group_by_pdb.get(pdb, "")
        sigs = sorted(motifs_by_pdb.get(pdb, ()),
                      key=lambda s: rank.get(s, 1e9))
        motif_cells = list(sigs[:n_motif_cols])
        if len(sigs) > n_motif_cols:            # don't drop extras
            motif_cells[-1] = "; ".join(sigs[n_motif_cols - 1:])
        motif_cells += [None] * (n_motif_cols - len(motif_cells))

        row = [grp, pdb]
        if do_enzyme:
            row += [best_by_pdb.get(pdb), cons_by_pdb.get(pdb)]
        row += motif_cells
        rows.append(row)
    return rows, motif_order


def _autofit(ws):
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None),
                default=0) + 2
        ws.column_dimensions[col[0].column_letter].width = min(w, 60)


def _add_sheet(wb, title, header, rows, first=False):
    from openpyxl.styles import Font
    ws = wb.active if first else wb.create_sheet(title)
    if first:
        ws.title = title
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    _autofit(ws)
    return ws


def write_excel(out, motif_summary, motif_detail,
                enzyme_summary, enzyme_detail, do_motif, do_enzyme,
                n_motif_cols=3):
    import openpyxl
    wb = openpyxl.Workbook()
    first = True

    # ---- overview: one row per PDB ID, joining both analyses ----
    overview_rows, _ = build_overview(motif_detail, enzyme_summary,
                                      do_motif, do_enzyme, n_motif_cols)
    ov_header = ["group", "pdb_id"]
    if do_enzyme:
        ov_header += ["best_hit_name", "consensus_name"]
    ov_header += [f"motif{i}" for i in range(1, n_motif_cols + 1)]
    _add_sheet(wb, "overview", ov_header, overview_rows, first=first)
    first = False

    if do_motif:
        _add_sheet(wb, "motif_summary",
                   ["group", "n_sequences", "ps_accession", "signature",
                    "n_seqs_with_motif", "pct_seqs"],
                   motif_summary, first=first)
        first = False
        _add_sheet(wb, "motif_hits",
                   ["group", "sequence_id", "ps_accession", "signature",
                    "start", "stop", "matched_substring"],
                   motif_detail)

    if do_enzyme:
        _add_sheet(wb, "enzyme_summary",
                   ["group", "query_entity", "query_len",
                    "best_hit_entity", "best_hit_name", "best_hit_ec",
                    "best_hit_uniprot", "best_hit_identity",
                    "best_hit_evalue", "best_hit_bitscore",
                    "consensus_name", "consensus_ec", "consensus_uniprot",
                    "consensus_ec_agreement", "n_hits_considered"],
                   enzyme_summary, first=first)
        first = False
        _add_sheet(wb, "enzyme_hits",
                   ["group", "query_entity", "hit_entity", "hit_kind",
                    "hit_name", "hit_ec", "hit_uniprot",
                    "identity", "evalue", "bitscore"],
                   enzyme_detail)

    wb.save(out)


# ========================================================================== #
# Main
# ========================================================================== #
def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input", help="CSV file of PDB IDs.")
    ap.add_argument("-o", "--out", default="pdb_motif_and_enzyme.xlsx",
                    help="Output Excel file.")
    # shared CSV / fetch options
    ap.add_argument("--id-col", default="pdb_id",
                    help="Column holding 4-char PDB IDs (default: pdb_id).")
    ap.add_argument("--group-col", default="cluster",
                    help="Column to group rows by (default: cluster).")
    ap.add_argument("--no-group", action="store_true",
                    help="Treat all IDs as a single group.")
    ap.add_argument("--all-chains", action="store_true",
                    help="Keep every chain/entity record (no per-entry dedup).")
    ap.add_argument("--keep-fasta", metavar="DIR",
                    help="Also write intermediate per-group .fasta files here.")
    # which analyses to run
    ap.add_argument("--skip-motif", action="store_true",
                    help="Do not run the PROSITE motif scan.")
    ap.add_argument("--skip-enzyme", action="store_true",
                    help="Do not run the RCSB enzyme-ID search.")
    # motif (ScanProsite) options
    ap.add_argument("--chunk", type=int, default=10,
                    help="Sequences per ScanProsite request (keep small).")
    ap.add_argument("--motif-sleep", type=float, default=4.0,
                    help="Seconds between ScanProsite requests (default 4).")
    ap.add_argument("--lowscore", action="store_true",
                    help="Include low-level (weak) PROSITE profile matches.")
    ap.add_argument("--prosite", help="Optional prosite.dat for signature names.")
    # enzyme (sequence search) options
    ap.add_argument("--identity", type=float, default=0.3,
                    help="MMseqs2 identity cutoff 0-1 (default 0.3).")
    ap.add_argument("--evalue", type=float, default=1.0,
                    help="E-value cutoff (default 1.0).")
    ap.add_argument("--rows", type=int, default=25,
                    help="Max hits to request per query (default 25).")
    ap.add_argument("--top", type=int, default=10,
                    help="Top hits to enrich + keep per query (default 10).")
    ap.add_argument("--enzyme-sleep", type=float, default=1.0,
                    help="Seconds between sequence-search requests (default 1).")
    ap.add_argument("--max-retries", type=int, default=3)
    args = ap.parse_args()

    do_motif = not args.skip_motif
    do_enzyme = not args.skip_enzyme
    if not (do_motif or do_enzyme):
        sys.exit("Nothing to do: both --skip-motif and --skip-enzyme set.")
    if not os.path.isfile(args.input):
        sys.exit(f"Input not found: {args.input}")

    groups = read_groups(args.input, args.id_col, args.group_col, args.no_group)
    if not groups:
        sys.exit("No PDB IDs found in input.")
    total_ids = sum(len(v) for v in groups.values())
    ran = " + ".join(x for x, on in [("motif", do_motif),
                                      ("enzyme-ID", do_enzyme)] if on)
    print(f"Read {len(groups)} group(s), {total_ids} unique ID(s) from "
          f"{args.input}  |  running: {ran}")

    names = load_prosite_names(args.prosite) if (do_motif and args.prosite) \
        else {}
    if args.keep_fasta:
        os.makedirs(args.keep_fasta, exist_ok=True)

    seq_cache = {}            # pid -> [(header, seq), ...]
    search_cache = {}         # seq -> hit list         (enzyme)
    enrich_cache = {}         # entity_id -> {name,ec}  (enzyme)

    motif_summary, motif_detail = [], []
    enzyme_summary, enzyme_detail = [], []

    for gi, (grp, ids) in enumerate(groups.items(), 1):
        print(f"\n=== group {grp}  ({gi}/{len(groups)}, {len(ids)} IDs) ===")

        # ---- fetch ONCE, shared by both analyses ----
        raw, got = fetch_ids(ids, seq_cache)
        rec_map = build_records(raw, args.all_chains)
        records = list(rec_map.values())          # list of (entity_id, seq)
        failed = [pid for pid in ids if pid not in got]
        print(f"  {len(records)} sequences"
              + (f"  ({len(failed)} IDs returned nothing)" if failed else ""))

        if args.keep_fasta:
            safe = "".join(ch if ch.isalnum() or ch in "-_." else "_"
                           for ch in str(grp))
            write_fasta_file(rec_map, os.path.join(args.keep_fasta,
                                                   f"group_{safe}.fasta"))

        if not records:
            if do_motif:
                motif_summary.append([str(grp), 0, "(none)", "(none)",
                                      0, "0.0%"])
            continue

        # ---- (A) motif ----
        if do_motif:
            matches = motif_scan_group(records, args.lowscore, args.chunk,
                                       args.max_retries, args.motif_sleep)
            ms, md = motif_summarize(str(grp), records, matches, names)
            motif_summary.extend(ms)
            motif_detail.extend(md)
            for row in ms:
                if row[2] != "(none)":
                    print(f"  [motif] {row[2]:10s} {row[3]:20s} {row[4]:4d} "
                          f"seqs ({row[5]})")
            if not md:
                print("  [motif] no PROSITE signatures matched")

        # ---- (B) enzyme ID ----
        if do_enzyme:
            es, ed = enzyme_id_group(
                str(grp), records, args.identity, args.evalue, args.rows,
                args.top, args.max_retries, args.enzyme_sleep,
                search_cache, enrich_cache)
            enzyme_summary.extend(es)
            enzyme_detail.extend(ed)

    write_excel(args.out, motif_summary, motif_detail,
                enzyme_summary, enzyme_detail, do_motif, do_enzyme)

    parts = []
    if do_motif:
        parts.append(f"motif: {len(motif_summary)} summary / "
                     f"{len(motif_detail)} hit rows")
    if do_enzyme:
        parts.append(f"enzyme: {len(enzyme_summary)} summary / "
                     f"{len(enzyme_detail)} hit rows")
    print(f"\nWrote {args.out}  ({'; '.join(parts)})")


if __name__ == "__main__":
    main()
